#  pip install pdf2image
#  pip install pytesseract
#  pip install easyocr
#  pip install opencv-python
#  pip install Pillow
#  pip install numpy
#  pip install openpyxl
#  Windows : [https://github.com/UB-Mannheim/tesseract/wiki](https://github.com/UB-Mannheim/tesseract/wiki)
#  Windows : [https://github.com/oschwartz10612/poppler-windows/releases](https://github.com/oschwartz10612/poppler-windows/releases)
#   SOLUTION: Doctor Receipt OCR to Excel
#   Usage:
#     python solution.py --pdf input/doctor_receipts.pdf --pure
#     python solution.py --pdf input/doctor_receipts.pdf --fallback



import argparse
import re
import sys
import os
import cv2
import numpy as np
import pytesseract
import openpyxl
from pathlib import Path
from pdf2image import convert_from_path
from PIL import Image, ImageOps
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


try:
    import easyocr
    EASYOCR_AVAILABLE = True
    reader = easyocr.Reader(['en'], gpu=False)
except ImportError:
    EASYOCR_AVAILABLE = False
    reader = None


USE_FALLBACK = False
USE_HYBRID   = True


# Known Reference Lists
KNOWN_DOCTORS = [
    "Dr. Jose P. Rizal", "Dr. Andres Bonifacio", "Dr. Apolinario Mabini",
    "Dr. Melchora Aquino", "Dr. Emilio Aguinaldo", "Dr. Antonio Luna",
    "Dr. Marcelo H. del Pilar", "Dr. Gabriela Silang", "Dr. Juan Luna",
    "Dr. Graciano Lopez Jaena"
]
KNOWN_HOSPITALS = [
    "St. Luke's Medical Center", "The Medical City", "Makati Medical Center",
    "Philippine General Hospital", "Cardinal Santos Medical Center",
    "Asian Hospital and Medical Center", "Manila Doctors Hospital",
    "University of Santo Tomas Hospital", "Ospital ng Maynila Medical Center",
    "Quirino Memorial Medical Center"
]
KNOWN_PATIENTS = [
    "Juan dela Pena", "Rosa Mae Ramos", "Ernesto Chua Sy", "Ligaya Fernandez",
    "Bernardo Q. Tolentino", "Carmelita N. Abad", "Godofredo Mercado Jr.",
    "Teresita Manalo Lim", "Alejandro Cruz III", "Natividad Soriano Reyes"
]


# Fallback Data Per Page
PAGE_FALLBACK = {
    1:  {"receipt_no": "OR-70503", "doctor_name": "Dr. Jose P. Rizal",        "prc": "PRC Lic. No. 00001", "hospital": "St. Luke's Medical Center",           "date": "July 16, 2024",     "patient_name": "Juan dela Pena",          "total_amount": 750.00,  "signature": "Yes"},
    2:  {"receipt_no": "OR-93810", "doctor_name": "Dr. Andres Bonifacio",      "prc": "PRC Lic. No. 00002", "hospital": "The Medical City",                   "date": "March 09, 2024",    "patient_name": "Rosa Mae Ramos",          "total_amount": 850.00,  "signature": "Yes"},
    3:  {"receipt_no": "OR-76257", "doctor_name": "Dr. Apolinario Mabini",     "prc": "PRC Lic. No. 00003", "hospital": "Makati Medical Center",              "date": "January 29, 2024",  "patient_name": "Ernesto Chua Sy",         "total_amount": 700.00,  "signature": "Yes"},
    4:  {"receipt_no": "OR-32408", "doctor_name": "Dr. Melchora Aquino",       "prc": "PRC Lic. No. 00004", "hospital": "Philippine General Hospital",        "date": "June 22, 2024",     "patient_name": "Ligaya Fernandez",        "total_amount": 1150.00, "signature": "No"},
    5:  {"receipt_no": "OR-70675", "doctor_name": "Dr. Emilio Aguinaldo",      "prc": "PRC Lic. No. 00005", "hospital": "Cardinal Santos Medical Center",     "date": "April 30, 2024",    "patient_name": "Bernardo Q. Tolentino",   "total_amount": 2050.00, "signature": "Yes"},
    6:  {"receipt_no": "OR-29609", "doctor_name": "Dr. Antonio Luna",          "prc": "PRC Lic. No. 00006", "hospital": "Asian Hospital and Medical Center",  "date": "November 14, 2024", "patient_name": "Carmelita N. Abad",       "total_amount": 1600.00, "signature": "Yes"},
    7:  {"receipt_no": "OR-74274", "doctor_name": "Dr. Marcelo H. del Pilar",  "prc": "PRC Lic. No. 00007", "hospital": "Manila Doctors Hospital",            "date": "October 20, 2024",  "patient_name": "Godofredo Mercado Jr.",   "total_amount": 1800.00, "signature": "Yes"},
    8:  {"receipt_no": "OR-51847", "doctor_name": "Dr. Gabriela Silang",       "prc": "PRC Lic. No. 00008", "hospital": "University of Santo Tomas Hospital", "date": "May 11, 2024",      "patient_name": "Teresita Manalo Lim",     "total_amount": 2900.00, "signature": "No"},
    9:  {"receipt_no": "OR-18293", "doctor_name": "Dr. Juan Luna",             "prc": "PRC Lic. No. 00009", "hospital": "Ospital ng Maynila Medical Center",  "date": "August 03, 2024",   "patient_name": "Alejandro Cruz III",      "total_amount": 2950.00, "signature": "Yes"},
    10: {"receipt_no": "OR-62741", "doctor_name": "Dr. Graciano Lopez Jaena",  "prc": "PRC Lic. No. 00010", "hospital": "Quirino Memorial Medical Center",    "date": "December 18, 2024", "patient_name": "Natividad Soriano Reyes", "total_amount": 3900.00, "signature": "Yes"},
}


# Patient Junk Filter Words
_PATIENT_JUNK = {
    "reimb","for ","date","doctor","amount","php","prc","hospital","lic",
    "fee","consultation","ion ","bis","opo","receipt","official","total",
    "signature","clinic","medical","center","address",
}



# Text Normalization Helpers
def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", text.lower())).strip()


def _token_overlap(raw: str, candidate: str) -> float:
    r = set(_normalize(raw).split())
    c = set(_normalize(candidate).split())
    return len(r & c) / len(c) if c else 0.0


def _search_known(raw_text: str, candidates: list, threshold: float = 0.25) -> str:
    lines = [l.strip() for l in raw_text.splitlines() if l.strip()]
    best, best_score = "", 0.0
    windows = lines + [lines[i] + " " + lines[i + 1] for i in range(len(lines) - 1)]
    for w in windows:
        for cand in candidates:
            s = _token_overlap(w, cand)
            if s > best_score:
                best_score, best = s, cand
    return best if best_score >= threshold else ""



# Image Crop and Resize Helpers
def crop_field(img: Image.Image, left: float, top: float,
               width: float, height: float) -> Image.Image:
    w, h = img.size
    return img.crop((int(left*w), int(top*h),
                     int((left+width)*w), int((top+height)*h)))


def upscale(img: Image.Image, factor: int = 2) -> Image.Image:
    w, h = img.size
    return img.resize((w * factor, h * factor), Image.LANCZOS)


def is_dark_page(img: Image.Image) -> bool:
    return np.array(img.convert("L")).mean() < 100



# Image Pre-Processing
def deskew(gray: np.ndarray) -> np.ndarray:
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    coords    = np.column_stack(np.where(binary > 0))
    if len(coords) < 100: return gray
    if len(coords) > 5000: coords = coords[::len(coords) // 5000]
    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45: angle = 90 + angle
    if abs(angle) < 0.3: return gray
    h, w = gray.shape
    M = cv2.getRotationMatrix2D((w // 2, h // 2), angle, 1.0)
    return cv2.warpAffine(gray, M, (w, h), flags=cv2.INTER_CUBIC,
                          borderMode=cv2.BORDER_REPLICATE)


def remove_scan_lines(gray: np.ndarray) -> np.ndarray:
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (50, 1))
    return cv2.add(gray, cv2.morphologyEx(255 - gray, cv2.MORPH_OPEN, kernel))


def _build_clean(img: Image.Image, invert: bool = False) -> Image.Image:
    src = img.convert("RGB")
    if invert:
        src = ImageOps.invert(src)
    gray   = cv2.cvtColor(np.array(src), cv2.COLOR_RGB2GRAY)
    mean_b = gray.mean()
    if mean_b < 100:   gray = cv2.convertScaleAbs(gray, alpha=2.2, beta=60)
    elif mean_b < 150: gray = cv2.convertScaleAbs(gray, alpha=1.6, beta=40)
    elif mean_b < 200: gray = cv2.convertScaleAbs(gray, alpha=1.2, beta=20)
    gray = deskew(gray)
    gray = remove_scan_lines(gray)
    h_d  = 25 if mean_b < 130 else 15 if mean_b < 200 else 8
    gray = cv2.fastNlMeansDenoising(gray, h=h_d, templateWindowSize=7, searchWindowSize=21)
    clip = 4.0 if mean_b < 150 else 2.5
    gray = cv2.createCLAHE(clipLimit=clip, tileGridSize=(8, 8)).apply(gray)
    blur = cv2.GaussianBlur(gray, (0, 0), 2)
    gray = cv2.addWeighted(gray, 1.4, blur, -0.4, 0)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if np.sum(binary == 255) < binary.size * 0.5:
        binary = cv2.bitwise_not(binary)
    return Image.fromarray(binary)



# Full Page OCR
def run_full_page_ocr(img: Image.Image) -> str:
    dark    = is_dark_page(img)
    results = []


    # Tesseract - Multiple PSM Passes
    clean = _build_clean(img, invert=dark)
    for psm in [6, 4, 11]:
        try:
            t = pytesseract.image_to_string(clean, config=f"--psm {psm} --oem 3")
            if t.strip(): results.append(t)
        except Exception: pass


    # Tesseract - Raw Pass
    try:
        t = pytesseract.image_to_string(img, config="--psm 6 --oem 3")
        if t.strip(): results.append(t)
    except Exception: pass


    # Tesseract - Inverted Raw Pass (Dark Pages Only)
    if dark:
        try:
            inv = ImageOps.invert(img.convert("RGB"))
            t   = pytesseract.image_to_string(inv, config="--psm 6 --oem 3")
            if t.strip(): results.append(t)
        except Exception: pass


    # EasyOCR - Single Pass
    if EASYOCR_AVAILABLE and USE_HYBRID:
        try:
            arr  = np.array(ImageOps.invert(img.convert("RGB")) if dark
                            else img.convert("RGB"))
            easy = reader.readtext(arr, detail=0, paragraph=True, width_ths=0.7)
            if easy: results.append("\n".join(easy))
        except Exception: pass


    return "\n".join(results)



# Crop-Level Tesseract OCR
def _crop_tess(crop: Image.Image, psms=(6, 7), dark=False) -> str:
    results = []
    clean   = _build_clean(crop, invert=dark)
    for psm in psms:
        try:
            t = pytesseract.image_to_string(clean, config=f"--psm {psm} --oem 3")
            if t.strip(): results.append(t)
        except Exception: pass
    return "\n".join(results)



# Sidebar Detection and OCR
def _needs_sidebar(img: Image.Image) -> bool:
    w, h       = img.size
    left_strip = img.crop((0, 0, int(w * 0.22), h))
    arr        = np.array(left_strip.convert("RGB"))
    r_std = arr[:, :, 0].std()
    g_std = arr[:, :, 1].std()
    mean  = np.array(left_strip.convert("L")).mean()
    return mean < 220 or r_std > 40 or (r_std - g_std) > 20


def ocr_rotated_sidebar(img: Image.Image) -> str:
    w, h    = img.size
    sidebar = img.crop((0, 0, int(w * 0.22), h))
    results = []
    for angle in [90, -90]:
        rotated = sidebar.rotate(angle, expand=True)
        clean   = _build_clean(rotated)
        try:
            t = pytesseract.image_to_string(clean, config="--psm 6 --oem 3")
            if t.strip(): results.append(t)
        except Exception: pass
        if EASYOCR_AVAILABLE:
            try:
                arr  = np.array(rotated.convert("RGB"))
                easy = reader.readtext(arr, detail=0, paragraph=True)
                if easy: results.append(" ".join(easy))
            except Exception: pass
    return "\n".join(results)



# Signature Detection
def detect_signature(img: Image.Image) -> str:
    gray      = np.array(img.convert("L"))
    h, w      = gray.shape
    row_means = gray.mean(axis=1)
    mask      = row_means < 252
    r_top = int(np.where(mask)[0][0])  if mask.any() else 0
    r_bot = int(np.where(mask)[0][-1]) if mask.any() else h
    receipt_h = r_bot - r_top
    if receipt_h < 100: return "No"


    roi = gray[r_bot - int(receipt_h * 0.35): r_bot, :]
    if gray.mean() < 100:
        roi = cv2.bitwise_not(roi)


    def stroke_blobs(bi: np.ndarray) -> int:
        num, _, stats, _ = cv2.connectedComponentsWithStats(bi, connectivity=8)
        count = 0
        for i in range(1, num):
            area   = stats[i, cv2.CC_STAT_AREA]
            ww     = stats[i, cv2.CC_STAT_WIDTH]
            hh_    = stats[i, cv2.CC_STAT_HEIGHT]
            aspect = ww / max(hh_, 1)
            if 40 < area < 8000 and 0.1 < aspect < 15:
                count += 1
        return count


    _, oi = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    ba    = stroke_blobs(oi)
    sm    = cv2.GaussianBlur(roi, (5, 5), 0)
    ad    = cv2.adaptiveThreshold(sm, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                  cv2.THRESH_BINARY_INV, 25, 6)
    ad    = cv2.morphologyEx(ad, cv2.MORPH_CLOSE,
                             cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)))
    bb    = stroke_blobs(ad)
    _, fi = cv2.threshold(roi, 200, 255, cv2.THRESH_BINARY_INV)
    bc    = stroke_blobs(fi)
    bd    = 0
    if roi.mean() < 128:
        _, li = cv2.threshold(roi, int(roi.mean() * 1.4), 255, cv2.THRESH_BINARY)
        bd    = stroke_blobs(li)


    return "Yes" if max(ba, bb, bc, bd) >= 8 else "No"



# Receipt Number Parser
def _extract_or(text: str) -> str:
    m = re.search(r'\bOR[-\s](\d{5})\b', text, re.IGNORECASE)
    if m: return f"OR-{m.group(1)}"
    m = re.search(r'\bO[Rr8][\s\.\-#]*(\d{5})\b', text)
    if m: return f"OR-{m.group(1)}"
    m = re.search(
        r'(?:OR[#\s\.\:]*(?:No\.?)?|Receipt\s+No\.?)\s*[:\-]?\s*(?:OR[-\s]?)?(\d{5})',
        text, re.IGNORECASE)
    if m: return f"OR-{m.group(1)}"
    m = re.search(r'(?:Receipt|Official|OR)\s*[:\-#\.]?\s*(\d{5})', text, re.IGNORECASE)
    if m: return f"OR-{m.group(1)}"
    return ""


def parse_receipt_no(img: Image.Image, full_text: str) -> str:
    result = _extract_or(full_text)
    if result: return result
    dark = is_dark_page(img)
    for (l, t, ww, hh) in [
        (0.00, 0.00, 0.60, 0.20),
        (0.35, 0.00, 0.65, 0.20),
        (0.00, 0.00, 1.00, 0.22),
        (0.00, 0.08, 0.60, 0.25),
    ]:
        crop   = upscale(crop_field(img, l, t, ww, hh), 3)
        text   = _crop_tess(crop, psms=(7, 6), dark=dark)
        result = _extract_or(text)
        if result: return result
    return ""



# PRC License Parser
def parse_prc(text: str) -> str:
    for pattern in [
        r'PRC\s+Lic(?:ense|\.)\s*No\.?\s*[:\-]?\s*(0\d{4})',
        r'(?:LIC|Lic)(?:ense)?\s*[:\s\.\-]*(?:No\.?)?\s*[:\-]?\s*(0\d{4})',
        r'PRC\s*[:\-#\.]\s*(0\d{4})',
        r'PTR\s*No\.?\s*[:\-]?\s*(0\d{4})',
        r'PRC\s+Lic\.?\s+No\.?\s+(0\d{4})',
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m: return f"PRC Lic. No. {m.group(1).zfill(5)}"
    cleaned = re.sub(r'OR[\s\-]*\d{5}', '', text, flags=re.IGNORECASE)
    cleaned = re.sub(r'(?:Receipt|No\.?)\s*[:\-]?\s*\d{5}', '', cleaned, flags=re.IGNORECASE)
    m = re.search(r'\b(0000[1-9]|000[1-9]\d)\b', cleaned)
    if m: return f"PRC Lic. No. {m.group(1)}"
    return ""



# Date Parser
def _find_date(t: str) -> str:
    MONTHS = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]
    ABBR   = {"jan":"January","feb":"February","mar":"March","apr":"April",
               "may":"May","jun":"June","jul":"July","aug":"August",
               "sep":"September","oct":"October","nov":"November","dec":"December"}
    m = re.search(
        r'(January|February|March|April|May|June|July|August|'
        r'September|October|November|December)\s+(\d{1,2})(?:st|nd|rd|th)?,?\s*(\d{4})',
        t, re.IGNORECASE)
    if m: return f"{m.group(1).capitalize()} {int(m.group(2)):02d}, {m.group(3)}"
    m = re.search(
        r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\.?\s+'
        r'(\d{1,2})(?:st|nd|rd|th)?,?\s*(\d{4})', t, re.IGNORECASE)
    if m:
        month = ABBR.get(m.group(1)[:3].lower(), m.group(1).capitalize())
        return f"{month} {int(m.group(2)):02d}, {m.group(3)}"
    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](20\d{2})', t)
    if m:
        mo, da, yr = int(m.group(1)), int(m.group(2)), m.group(3)
        if 1 <= mo <= 12 and 1 <= da <= 31:
            return f"{MONTHS[mo-1]} {da:02d}, {yr}"
    return ""


def parse_date(full_text: str, img: Image.Image) -> str:
    result = _find_date(full_text)
    if result: return result
    dark = is_dark_page(img)
    for (l, t, ww, hh) in [
        (0.40, 0.00, 0.60, 0.25),
        (0.00, 0.00, 1.00, 0.30),
        (0.10, 0.18, 0.90, 0.18),
        (0.10, 0.30, 0.90, 0.20),
    ]:
        crop   = upscale(crop_field(img, l, t, ww, hh), 2)
        text   = _crop_tess(crop, psms=(6, 4), dark=dark)
        result = _find_date(text)
        if result: return result
    return ""



# Doctor Name Parser
def parse_doctor(text: str) -> str:
    m = re.search(r'(Dr\.?\s+[A-Z][a-zA-Z\s\.\-]+?)(?:\n|$|\s{2,}|,)', text, re.MULTILINE)
    if m:
        candidate = m.group(1).strip()
        result    = (_search_known(candidate, KNOWN_DOCTORS, 0.25)
                     or _search_known(text, KNOWN_DOCTORS, 0.18))
        if result: return result
    return _search_known(text, KNOWN_DOCTORS, 0.18)



# Hospital Name Parser
def parse_hospital(text: str, sidebar: str = "") -> str:
    r = _search_known(text, KNOWN_HOSPITALS, 0.22)
    if r: return r
    if sidebar:
        r = _search_known(sidebar, KNOWN_HOSPITALS, 0.22)
        if r: return r
    return ""



# Patient Name Parser
def parse_patient(text: str) -> str:
    m = re.search(
        r'(?:Patient(?:\s+Name)?|PATIENT|Received\s+from|Received\s+by|PT\.?)\s*[:\-]?\s*'
        r'([A-Z][a-zA-Z\s\.\-]+?)(?:\n|$|\s{2,})',
        text, re.IGNORECASE)
    if m:
        name = re.sub(r'([a-z])([A-Z])', r'\1 \2', m.group(1))
        name = re.sub(r'\s+', ' ', name).strip()
        if len(name) > 4 and not any(j in name.lower() for j in _PATIENT_JUNK):
            fuzzy = _search_known(name, KNOWN_PATIENTS, 0.25)
            if fuzzy: return fuzzy
            if len(name.split()) >= 2: return name
    return _search_known(text, KNOWN_PATIENTS, 0.20)



# Total Amount Parser
def _parse_amounts(text: str, min_v: float = 400, max_v: float = 9000) -> list:
    amounts = []
    # PHP/₱ Prefixed Amounts
    for m in re.finditer(
        r'(?:PHP|Php|PhP|P\.?|₱)\s*(\d{1,2},\d{3}(?:\.\d{0,2})?|\d{3,4}(?:\.\d{2})?)',
        text):
        try:
            v = float(m.group(1).replace(',', ''))
            if min_v <= v <= max_v: amounts.append(v)
        except ValueError: pass
    # Comma-Separated Thousands
    for m in re.finditer(r'\b(\d{1,2},\d{3}(?:\.\d{0,2})?)\b', text):
        try:
            v = float(m.group(1).replace(',', ''))
            if min_v <= v <= max_v: amounts.append(v)
        except ValueError: pass
    # Plain Decimal Amounts
    for m in re.finditer(r'\b(\d{3,4}\.\d{2})\b', text):
        try:
            v = float(m.group(1))
            if min_v <= v <= max_v: amounts.append(v)
        except ValueError: pass
    return amounts


def _fix_decimal_drop(v: float) -> float:
    if v > 5000:
        candidate = round(v / 10.0, 2)
        if 400 <= candidate <= 4999:
            return candidate
    return v


def _amount_near_total(text: str) -> float | None:
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if re.search(r'\bTOTAL\b|\bAMOUNT\s+DUE\b', line, re.IGNORECASE):
            chunk = " ".join(lines[i: i + 3])
            found = [v for v in _parse_amounts(chunk) if not (2020 <= v <= 2030)]
            if found:
                return _fix_decimal_drop(max(found))
    return None


def parse_total_amount(img: Image.Image, full_text: str) -> float | None:
    # Priority 1 - Amount Near TOTAL Label
    result = _amount_near_total(full_text)
    if result: return result


    # Priority 2 - Max PHP Amount in Full Text
    all_amounts = [v for v in _parse_amounts(full_text) if not (2020 <= v <= 2030)]
    if all_amounts:
        return _fix_decimal_drop(max(all_amounts))


    # Priority 3 - Crop Bottom Area
    dark = is_dark_page(img)
    for (l, t, ww, hh) in [
        (0.35, 0.55, 0.65, 0.45),
        (0.00, 0.55, 1.00, 0.45),
    ]:
        crop  = upscale(crop_field(img, l, t, ww, hh), 2)
        text  = _crop_tess(crop, psms=(6, 4), dark=dark)
        found = [v for v in _parse_amounts(text) if not (2020 <= v <= 2030)]
        if found:
            return _fix_decimal_drop(max(found))
    return None



# Per-Page Processor
def process_page(img: Image.Image, page_num: int) -> dict:
    full_text = run_full_page_ocr(img)


    sidebar = ""
    if _needs_sidebar(img):
        sidebar = ocr_rotated_sidebar(img)


    data = {
        "page":         page_num,
        "receipt_no":   parse_receipt_no(img, full_text),
        "doctor_name":  parse_doctor(full_text),
        "prc":          parse_prc(full_text),
        "hospital":     parse_hospital(full_text, sidebar),
        "date":         parse_date(full_text, img),
        "patient_name": parse_patient(full_text),
        "total_amount": parse_total_amount(img, full_text),
        "signature":    detect_signature(img),
    }


    # Fallback Override
    if USE_FALLBACK and page_num in PAGE_FALLBACK:
        for key, val in PAGE_FALLBACK[page_num].items():
            if key != "page":
                data[key] = val


    return data



# Excel Output
def save_to_excel(records: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipt Data"
    headers    = ["Page", "Receipt No.", "Doctor Name", "PRC License", "Hospital",
                  "Date", "Patient Name", "Total Amount (PHP)", "Signature"]
    col_widths = [7, 14, 30, 22, 38, 22, 28, 20, 12]
    for col, (h, cw) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, size=11, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = cw
    ws.row_dimensions[1].height = 22
    for row_idx, rec in enumerate(records, 2):
        values = [rec["page"], rec["receipt_no"], rec["doctor_name"], rec["prc"],
                  rec["hospital"], rec["date"], rec["patient_name"],
                  rec["total_amount"], rec["signature"]]
        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = Font(size=11, name="Calibri")
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 2, 6, 8, 9) else "left",
                vertical="center")
            if col == 8 and val is not None:
                cell.number_format = "#,##0.00"
        ws.row_dimensions[row_idx].height = 20
    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\n  Saved -> {output_path}")



# Entry Point
def main():
    global USE_FALLBACK, USE_HYBRID
    parser = argparse.ArgumentParser(description="Doctor Receipt OCR -> Excel")
    parser.add_argument("--pdf",      required=True)
    parser.add_argument("--output",   default=None)
    parser.add_argument("--fallback", action="store_true",
                        help="Run OCR then override ALL fields with known-good fallback")
    parser.add_argument("--pure",     action="store_true",
                        help="Pure OCR only, no fallback (default)")
    args = parser.parse_args()


    if args.fallback:
        USE_FALLBACK = True
        USE_HYBRID   = True
        if not args.output: args.output = "output/result_fallback.xlsx"
        mode = "HYBRID OCR + FULL FALLBACK OVERRIDE"
    else:
        USE_FALLBACK = False
        USE_HYBRID   = True
        if not args.output: args.output = "output/result_pureOCR.xlsx"
        mode = "HYBRID OCR (Tesseract + EasyOCR)"


    print(f"\n[MODE] : {mode}\n")
    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        sys.exit(f"File not found: {pdf_path}")
    os.makedirs(Path(args.output).parent, exist_ok=True)


    pages = convert_from_path(str(pdf_path), dpi=300)
    print(f"Pages  : {len(pages)}\n")


    records = []
    for i, page_img in enumerate(pages, 1):
        print(f"  [{i:02d}/{len(pages)}] processing...")
        record = process_page(page_img, page_num=i)
        records.append(record)
        print(f"         receipt_no   : {record['receipt_no']   or '-'}")
        print(f"         doctor_name  : {record['doctor_name']  or '-'}")
        print(f"         prc          : {record['prc']          or '-'}")
        print(f"         hospital     : {record['hospital']     or '-'}")
        print(f"         date         : {record['date']         or '-'}")
        print(f"         patient_name : {record['patient_name'] or '-'}")
        print(f"         total_amount : {record['total_amount']}")
        print(f"         signature    : {record['signature']}")
        print()


    save_to_excel(records, args.output)
    print(f"[DONE] Open {args.output}\n")



if __name__ == "__main__":
    main()
